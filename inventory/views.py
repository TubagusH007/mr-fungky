import re
import os
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction, DatabaseError
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.conf import settings
from django.utils import timezone
from .models import BahanBaku, RiwayatStok, DataHistorisHarian, BarangMasuk, Profile, LoginHistory, CatatanManager
from gtts import gTTS
from django.http import HttpResponse
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    Workbook = None

# 1. Halaman Utama (Dashboard)
@login_required
def home(request):
    # Ambil 10 riwayat update terakhir (urutkan berdasarkan waktu terbaru)
    history = RiwayatStok.objects.all().order_by('-waktu')[:10]
    
    # Ambil daftar bahan yang stoknya di bawah atau sama dengan stok minimal
    semua_bahan = BahanBaku.objects.all()
    menipis_teh = [b for b in semua_bahan if b.stok_sekarang <= b.stok_minimal and b.kategori_zona == 'TEA']
    menipis_bahan = [b for b in semua_bahan if b.stok_sekarang <= b.stok_minimal and b.kategori_zona in ['BAHAN_BAKU', 'DAIRY']]
    menipis_topping = [b for b in semua_bahan if b.stok_sekarang <= b.stok_minimal and b.kategori_zona == 'TOPPING']
    
    context = {
        'history': history,
        'menipis_teh': menipis_teh,
        'menipis_bahan': menipis_bahan,
        'menipis_topping': menipis_topping,
        'total_menipis': len(menipis_teh) + len(menipis_bahan) + len(menipis_topping)
    }
    
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if profile.role == 'MANAGER':
        context['riwayat_login'] = LoginHistory.objects.all().order_by('-waktu')[:10]
        context['catatan_manager'] = CatatanManager.objects.all().order_by('-waktu')[:10]
        
    return render(request, 'inventory/index.html', context)

@login_required
def hapus_riwayat(request, log_id):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'MANAGER':
        return JsonResponse({'status': 'gagal', 'pesan': 'Akses ditolak: Manager tidak dapat menghapus riwayat.'})
        
    if request.method == 'POST':
        try:
            riwayat = RiwayatStok.objects.get(id=log_id)
            riwayat.delete()
            return JsonResponse({'status': 'sukses'})
        except RiwayatStok.DoesNotExist:
            return JsonResponse({'status': 'gagal', 'pesan': 'Riwayat tidak ditemukan.'})
    return JsonResponse({'status': 'gagal', 'pesan': 'Method not allowed'})

# 2. Otak Parser Suara (Fuzzy Matching untuk Nama Panjang)
@login_required
def proses_suara_massal(request):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'MANAGER':
        return JsonResponse({'status': 'gagal', 'pesan': 'Akses ditolak: Manager tidak dapat memodifikasi stok.'})
        
    teks_input = request.GET.get('pesan', '').lower().strip()
    # Urutkan berdasarkan panjang nama (descending) agar nama yang lebih spesifik dicek duluan
    daftar_bahan = sorted(BahanBaku.objects.all(), key=lambda x: len(x.nama_bahan), reverse=True)
    hasil_update = []
    
    # Normalisasi variasi kata unit dan penyebutan bahan agar lebih konsisten
    replacements = {
        'derigen': 'drigen', 'jerigen': 'drigen', 
        'zak': 'sak', 'pack': 'pak',
        'dipros': 'deeproast', 'di pros': 'deeproast',
        'crimmer': 'creamer', 'crimir': 'creamer', 'krimer': 'creamer', 'crimer': 'creamer', 'cerimer': 'creamer', 'kerimer': 'creamer',
        'konjak': 'konjac', 'jeli': 'jelly',
        'nol': '0', 'satu': '1', 'dua': '2', 'tiga': '3', 'empat': '4', 'lima': '5',
        'enam': '6', 'tujuh': '7', 'delapan': '8', 'sembilan': '9', 'sepuluh': '10'
    }

    for old, new in replacements.items():
        teks_input = teks_input.replace(old, new)
    
    # 2. Otak Parser Suara (Multi-Item & Fuzzy Matching)
    # Pisahkan input berdasarkan pemisah umum (dan, lalu, kemudian, koma, titik, serta, juga)
    pemisah = [r'\bdan\b', r'\blalu\b', r'\bkemudian\b', r',', r'\.', r'\bserta\b', r'\bjuga\b']
    pattern_pemisah = '|'.join(pemisah)
    segmen_input = re.split(pattern_pemisah, teks_input)
    
    hasil_update = []
    last_action = 'set'
    
    # Kita looping setiap segmen hasil pemisahan
    for segmen in segmen_input:
        current_segmen = segmen.strip()
        if not current_segmen:
            continue
            
        if any(kata in current_segmen for kata in ['tambah', 'masuk', 'restok', 'plus']):
            last_action = 'tambah'
        elif any(kata in current_segmen for kata in ['terpakai', 'kurang', 'minus', 'keluar', 'dipakai']):
            last_action = 'kurang'
        elif any(kata in current_segmen for kata in ['sisa', 'jadi', 'menjadi']):
            last_action = 'set'
        if not current_segmen:
            continue
            
        # Untuk setiap segmen, kita cari bahan yang cocok
        # Kita looping daftar_bahan yang sudah diurutkan berdasarkan panjang nama
        for bahan in daftar_bahan:
            nama_db = bahan.nama_bahan.lower().strip()
            kata_kunci = nama_db.split()[0]
            
            # Cek apakah bahan ada di segmen ini
            is_match = nama_db in current_segmen or (kata_kunci in current_segmen and len(kata_kunci) > 3 and kata_kunci not in ['gula', 'paper', 'cup'])
            
            # Penanganan khusus Gula Cair jika ada kata 'drigen' di segmen
            if 'drigen' in current_segmen and 'gula cair' in nama_db:
                is_match = True
                
            if is_match:
                # Cari angka di segmen ini
                match_angka = re.search(r"(\d+[,.]?\d*)", current_segmen)
                
                if match_angka:
                    angka_raw = float(match_angka.group(1).replace(',', '.'))
                    angka_final = angka_raw
                    
                    # Logika khusus Gula Cair dengan satuan Drigen (1 Drigen = 10.000 ml)
                    info_unit = f" {bahan.satuan}"
                    if 'drigen' in current_segmen and 'gula cair' in nama_db:
                        angka_final = angka_raw * 10000
                        info_unit = f" ({int(angka_raw)} drigen)"
                    
                    # Logika khusus Creamer dengan satuan Sak (1 Sak = 25 kg)
                    elif 'sak' in current_segmen and 'creamer' in nama_db:
                        angka_final = angka_raw * 25
                        info_unit = f" ({int(angka_raw)} sak)"
                    
                    # Terapkan berdasarkan last_action (tambah, kurang, atau set)
                    if last_action == 'tambah':
                        stok_akhir = bahan.stok_sekarang + angka_final
                        hasil_update.append(f"{bahan.nama_bahan} ditambah {int(angka_final)}{info_unit}")
                    elif last_action == 'kurang':
                        stok_akhir = max(0, bahan.stok_sekarang - angka_final)
                        hasil_update.append(f"{bahan.nama_bahan} dikurangi {int(angka_final)}{info_unit} sisa {int(stok_akhir)}")
                    else:
                        stok_akhir = angka_final
                        hasil_update.append(f"{bahan.nama_bahan} sisa {int(angka_final)}{info_unit}")
                    
                    # Update Stok di Database
                    bahan.stok_sekarang = stok_akhir
                    bahan.save()
                    
                    # Catat di RiwayatStok
                    RiwayatStok.objects.create(
                        petugas=request.user,
                        bahan=bahan,
                        jumlah_baru=stok_akhir
                    )
                    
                    # Masking: Hapus bahan dan angka yang sudah diproses dari current_segmen 
                    # agar tidak memicu deteksi bahan lain yang mirip atau angka yang sama di segmen yang sama
                    # Kita ganti dengan spasi agar posisi karakter lain tidak berubah secara drastis (meskipun di sini kita tidak pakai index setelahnya)
                    match_nama_index = current_segmen.find(nama_db)
                    if match_nama_index == -1 and kata_kunci in current_segmen:
                        match_nama_index = current_segmen.find(kata_kunci)
                    
                    if match_nama_index != -1:
                        len_nama = len(nama_db) if nama_db in current_segmen else len(kata_kunci)
                        current_segmen = current_segmen[:match_nama_index] + (" " * len_nama) + current_segmen[match_nama_index + len_nama:]
                    
                    start_a, end_a = match_angka.span()
                    current_segmen = current_segmen[:start_a] + (" " * (end_a - start_a)) + current_segmen[end_a:]
                    
                    # Kita tetap lanjut looping daftar_bahan untuk mencari bahan lain di segmen yang sama
                    # jika segmen tersebut ternyata mengandung beberapa bahan (misal tanpa pemisah 'dan')

    # 3. Logika Suara Konfirmasi (gTTS)
    if hasil_update:
        # Jika bahan yang diupdate banyak, ringkas pesan suara agar tidak terlalu panjang
        if len(hasil_update) > 3:
            teks_suara = f"Berhasil. {request.user.username}, {len(hasil_update)} bahan sudah diperbarui."
        else:
            teks_suara = f"Berhasil. {request.user.username}, " + " dan ".join(hasil_update) + " sudah diperbarui."
        
        tts = gTTS(text=teks_suara, lang='id')
        audio_path = os.path.join(settings.BASE_DIR, "static/konfirmasi_stok.mp3")
        
        try:
            if not os.path.exists(os.path.dirname(audio_path)):
                os.makedirs(os.path.dirname(audio_path))
            tts.save(audio_path)
        except Exception as e:
            print(f"Warning: gTTS Error - {e}")
    else:
        teks_suara = "Maaf Bee, sistem tidak mengenali bahan tersebut."

    return JsonResponse({
        'status': 'sukses' if hasil_update else 'gagal',
        'log': hasil_update,
        'pesan_suara': teks_suara
    })


@login_required
def halaman_prediksi(request):
    # Bersihkan data lama jika diperlukan untuk sinkronisasi Excel
    if not DataHistorisHarian.objects.filter(jumlah_antrian=450, honey_syrup=1).exists():
        DataHistorisHarian.objects.all().delete()
        # Data dari Excel Perkiraan Mr Fungky (Rasio Gula Cair = 80ml/antrian)
        DataHistorisHarian.objects.create(jumlah_antrian=300, omset=17000000, deeproast=1500, four_season=1350, jasmine=1050, boba=900, paper_cup_besar=600, paper_cup_kecil=600, gula=450, gula_cair=24000, honey_syrup=1, konjac_jelly=2, creamer=15)
        DataHistorisHarian.objects.create(jumlah_antrian=320, omset=17500000, deeproast=1650, four_season=1350, jasmine=1050, boba=1200, paper_cup_besar=640, paper_cup_kecil=640, gula=600, gula_cair=25600, honey_syrup=1, konjac_jelly=2, creamer=16)
        DataHistorisHarian.objects.create(jumlah_antrian=302, omset=16800000, deeproast=1650, four_season=1350, jasmine=1050, boba=1200, paper_cup_besar=604, paper_cup_kecil=604, gula=600, gula_cair=24160, honey_syrup=1, konjac_jelly=2, creamer=15.1)
        DataHistorisHarian.objects.create(jumlah_antrian=360, omset=18000000, deeproast=1800, four_season=1350, jasmine=1050, boba=1200, paper_cup_besar=720, paper_cup_kecil=720, gula=600, gula_cair=28800, honey_syrup=1, konjac_jelly=3, creamer=18)
        DataHistorisHarian.objects.create(jumlah_antrian=450, omset=22800000, deeproast=2100, four_season=1500, jasmine=1050, boba=2000, paper_cup_besar=900, paper_cup_kecil=900, gula=1000, gula_cair=36000, honey_syrup=1, konjac_jelly=3, creamer=50)
        DataHistorisHarian.objects.create(jumlah_antrian=567, omset=28000000, deeproast=2700, four_season=1800, jasmine=1350, boba=2300, paper_cup_besar=1134, paper_cup_kecil=1134, gula=1150, gula_cair=45360, honey_syrup=2, konjac_jelly=4, creamer=65)
        DataHistorisHarian.objects.create(jumlah_antrian=550, omset=27500000, deeproast=2550, four_season=1650, jasmine=1350, boba=2300, paper_cup_besar=1100, paper_cup_kecil=1100, gula=1150, gula_cair=44000, honey_syrup=2, konjac_jelly=4, creamer=60)
    
    data_historis = DataHistorisHarian.objects.all().order_by('-tanggal')[:10]
    return render(request, 'inventory/prediksi.html', {'data_historis': data_historis})

@login_required
def simpan_data_historis(request):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'MANAGER':
        return JsonResponse({'status': 'gagal', 'pesan': 'Akses ditolak: Manager tidak dapat memodifikasi data historis.'})
        
    if request.method == 'POST':
        try:
            DataHistorisHarian.objects.create(
                jumlah_antrian=int(request.POST.get('jumlah_antrian')),
                omset=float(request.POST.get('omset')),
                deeproast=float(request.POST.get('deeproast')),
                four_season=float(request.POST.get('four_season')),
                jasmine=float(request.POST.get('jasmine')),
                boba=float(request.POST.get('boba')),
                paper_cup_besar=int(request.POST.get('paper_cup_besar')),
                paper_cup_kecil=int(request.POST.get('paper_cup_kecil')),
                gula=float(request.POST.get('gula')),
                gula_cair=float(request.POST.get('gula_cair')),
                honey_syrup=float(request.POST.get('honey_syrup')),
                konjac_jelly=float(request.POST.get('konjac_jelly')),
                creamer=float(request.POST.get('creamer', 0))
            )
            return JsonResponse({'status': 'sukses'})
        except Exception as e:
            return JsonResponse({'status': 'gagal', 'pesan': str(e)})
    return JsonResponse({'status': 'gagal', 'pesan': 'Method not allowed'})

@login_required
def api_kalkulasi_prediksi(request):
    try:
        target_antrian = int(request.GET.get('antrian', 0))
        target_omset = float(request.GET.get('omset', 0))
        
        semua_data = DataHistorisHarian.objects.all()
        if not semua_data.exists() or target_antrian == 0:
            return JsonResponse({'status': 'gagal', 'pesan': 'Data latih tidak cukup atau antrian 0'})
            
        total_antrian = sum([d.jumlah_antrian for d in semua_data])
        
        # Hitung rasio rata-rata bahan baku per 1 antrian
        # Hanya hitung dari data yang > 0 agar tidak terdelusi oleh data lama/kosong
        def hitung_rasio(field_name):
            data_valid = [getattr(d, field_name) for d in semua_data if getattr(d, field_name) > 0]
            antrian_valid = [d.jumlah_antrian for d in semua_data if getattr(d, field_name) > 0]
            if not data_valid: return 0
            return sum(data_valid) / sum(antrian_valid)

        rasio = {
            'deeproast': hitung_rasio('deeproast'),
            'four_season': hitung_rasio('four_season'),
            'jasmine': hitung_rasio('jasmine'),
            'boba': hitung_rasio('boba'),
            'paper_cup_besar': hitung_rasio('paper_cup_besar'),
            'paper_cup_kecil': hitung_rasio('paper_cup_kecil'),
            'gula': hitung_rasio('gula'),
            'gula_cair': hitung_rasio('gula_cair'),
            'honey_syrup': hitung_rasio('honey_syrup'),
            'konjac_jelly': hitung_rasio('konjac_jelly'),
            'creamer': hitung_rasio('creamer'),
        }
        
        # Kalkulasi estimasi untuk hari ini berdasarkan target antrian
        prediksi = {k: round(v * target_antrian, 2) for k, v in rasio.items()}
        
        # Genapkan Konjac Jelly dan Cup (karena per pack/pcs, tidak ada koma)
        if 'konjac_jelly' in prediksi:
            prediksi['konjac_jelly'] = int(round(prediksi['konjac_jelly']))
        if 'paper_cup_besar' in prediksi:
            prediksi['paper_cup_besar'] = int(round(prediksi['paper_cup_besar']))
        if 'paper_cup_kecil' in prediksi:
            prediksi['paper_cup_kecil'] = int(round(prediksi['paper_cup_kecil']))

        
        return JsonResponse({'status': 'sukses', 'prediksi': prediksi})
    except Exception as e:
        return JsonResponse({'status': 'gagal', 'pesan': str(e)})

@login_required
def halaman_gudang(request):
    semua_bahan = BahanBaku.objects.all().order_by('kategori_zona', 'nama_bahan')
    riwayat_masuk = BarangMasuk.objects.all()[:10]
    return render(request, 'inventory/gudang.html', {
        'semua_bahan': semua_bahan,
        'riwayat_masuk': riwayat_masuk
    })

@login_required
def simpan_barang_masuk(request):
    if request.method == 'POST':
        bahan_id = request.POST.get('bahan_id')
        jumlah = request.POST.get('jumlah')
        penerima = request.POST.get('penerima')
        
        try:
            bahan = BahanBaku.objects.get(id=bahan_id)
            jumlah_float = float(jumlah)
            
            # Update stok utama
            bahan.stok_sekarang += jumlah_float
            bahan.save()
            
            # Catat riwayat barang masuk
            BarangMasuk.objects.create(
                bahan=bahan,
                jumlah=jumlah_float,
                penerima=penerima,
                zona=bahan.get_kategori_zona_display(),
                penginput=request.user
            )
            
            # Juga catat di RiwayatStok umum
            RiwayatStok.objects.create(
                petugas=request.user,
                bahan=bahan,
                jumlah_baru=bahan.stok_sekarang
            )
            
            return JsonResponse({'status': 'sukses', 'pesan': f'Stok {bahan.nama_bahan} berhasil ditambah.'})
        except Exception as e:
            return JsonResponse({'status': 'gagal', 'pesan': str(e)})
    return JsonResponse({'status': 'gagal', 'pesan': 'Method not allowed'})


@login_required
def update_stok_manual(request):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'MANAGER':
        return JsonResponse({'status': 'gagal', 'pesan': 'Akses ditolak: Manager tidak dapat memodifikasi stok.'})
        
    if request.method == 'POST':
        bahan_id = request.POST.get('bahan_id')
        stok_baru = request.POST.get('stok_baru')
        tipe_update = request.POST.get('tipe_update', 'set')
        
        try:
            bahan = BahanBaku.objects.get(id=bahan_id)
            angka_input = float(stok_baru)
            
            if tipe_update == 'set_gr':
                if bahan.satuan.lower() == 'kg':
                    angka_akhir = angka_input / 1000
                else:
                    angka_akhir = angka_input
            elif tipe_update == 'set_pack':
                # Konversi 1 pack = 20 cup (Menambah stok)
                angka_akhir = bahan.stok_sekarang + (angka_input * 20)
            elif tipe_update == 'tambah_kg':
                angka_akhir = bahan.stok_sekarang + angka_input
            else:
                angka_akhir = angka_input



                
            bahan.stok_sekarang = angka_akhir
            bahan.save()
            
            RiwayatStok.objects.create(
                petugas=request.user,
                bahan=bahan,
                jumlah_baru=angka_akhir
            )
            return JsonResponse({'status': 'sukses', 'pesan': f'{bahan.nama_bahan} berhasil diperbarui.'})
        except BahanBaku.DoesNotExist:
            return JsonResponse({'status': 'gagal', 'pesan': 'Bahan tidak ditemukan.'})
        except ValueError:
            return JsonResponse({'status': 'gagal', 'pesan': 'Format angka tidak valid.'})
        except Exception as e:
            return JsonResponse({'status': 'gagal', 'pesan': str(e)})
            
    return JsonResponse({'status': 'gagal', 'pesan': 'Method not allowed'})

@login_required
def terapkan_prediksi_stok(request):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'MANAGER':
        return JsonResponse({'status': 'gagal', 'pesan': 'Akses ditolak: Manager tidak dapat melakukan pemotongan stok prediksi.'})
        
    if request.method == 'POST':
        try:
            target_antrian = int(request.POST.get('antrian', 0))
            
            semua_data = DataHistorisHarian.objects.all()
            if not semua_data.exists() or target_antrian <= 0:
                return JsonResponse({'status': 'gagal', 'pesan': 'Data latih tidak cukup atau antrian tidak valid.'})
                
            total_antrian = sum([d.jumlah_antrian for d in semua_data])
            
            def hitung_rasio(field_name):
                data_valid = [getattr(d, field_name) for d in semua_data if getattr(d, field_name) > 0]
                antrian_valid = [d.jumlah_antrian for d in semua_data if getattr(d, field_name) > 0]
                if not data_valid: return 0
                return sum(data_valid) / sum(antrian_valid)

            rasio = {
                'deeproast': hitung_rasio('deeproast'),
                'four_season': hitung_rasio('four_season'),
                'jasmine': hitung_rasio('jasmine'),
                'boba': hitung_rasio('boba'),
                'paper_cup_besar': hitung_rasio('paper_cup_besar'),
                'paper_cup_kecil': hitung_rasio('paper_cup_kecil'),
                'gula': hitung_rasio('gula'),
                'gula_cair': hitung_rasio('gula_cair'),
                'honey_syrup': hitung_rasio('honey_syrup'),
                'konjac_jelly': hitung_rasio('konjac_jelly'),
                'creamer': hitung_rasio('creamer'),
            }
            
            prediksi = {k: round(v * target_antrian, 2) for k, v in rasio.items()}
            # Genapkan Konjac Jelly dan Cup (karena per pack/pcs, tidak ada koma)
            if 'konjac_jelly' in prediksi:
                prediksi['konjac_jelly'] = int(round(prediksi['konjac_jelly']))
            if 'paper_cup_besar' in prediksi:
                prediksi['paper_cup_besar'] = int(round(prediksi['paper_cup_besar']))
            if 'paper_cup_kecil' in prediksi:
                prediksi['paper_cup_kecil'] = int(round(prediksi['paper_cup_kecil']))

            
            daftar_bahan = BahanBaku.objects.all()
            bahan_diperbarui = []
            
            for bahan in daftar_bahan:
                nama_db = bahan.nama_bahan.lower()
                jumlah_potong = 0
                
                if 'deeproast' in nama_db: jumlah_potong = prediksi['deeproast']
                elif 'four season' in nama_db: jumlah_potong = prediksi['four_season']
                elif 'jasmine' in nama_db: jumlah_potong = prediksi['jasmine']
                elif 'boba' in nama_db: jumlah_potong = prediksi['boba']
                elif 'besar' in nama_db and 'cup' in nama_db: jumlah_potong = prediksi['paper_cup_besar']
                elif 'kecil' in nama_db and 'cup' in nama_db: jumlah_potong = prediksi['paper_cup_kecil']
                elif 'cair' in nama_db and 'gula' in nama_db: jumlah_potong = prediksi['gula_cair']
                elif 'gula' in nama_db and 'cair' not in nama_db: jumlah_potong = prediksi['gula']
                elif 'honey' in nama_db: jumlah_potong = prediksi['honey_syrup']
                elif 'konjac' in nama_db or 'jelly' in nama_db: jumlah_potong = prediksi['konjac_jelly']
                elif 'creamer' in nama_db: jumlah_potong = prediksi['creamer']
                
                if jumlah_potong > 0:
                    stok_baru = max(0, bahan.stok_sekarang - jumlah_potong) # Jangan sampai minus
                    bahan.stok_sekarang = stok_baru
                    bahan.save()
                    
                    RiwayatStok.objects.create(
                        petugas=request.user,
                        bahan=bahan,
                        jumlah_baru=stok_baru
                    )
                    bahan_diperbarui.append(bahan.nama_bahan)
            
            return JsonResponse({
                'status': 'sukses', 
                'pesan': 'Stok berhasil dipotong sesuai prediksi.',
                'bahan_diperbarui': bahan_diperbarui
            })
        except Exception as e:
            return JsonResponse({'status': 'gagal', 'pesan': str(e)})
            
    return JsonResponse({'status': 'gagal', 'pesan': 'Method not allowed'})

# --- Autentikasi ---

def halaman_login(request):
    if request.user.is_authenticated:
        return redirect('/')
        
    pesan_error = None
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        role = request.POST.get('role', 'KARYAWAN')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            profile, created = Profile.objects.get_or_create(user=user, defaults={'role': 'KARYAWAN'})
            if not user.is_superuser and profile.role != role:
                role_display = "Manager" if role == 'MANAGER' else "Karyawan"
                profile_role_display = "Manager" if profile.role == 'MANAGER' else "Karyawan"
                pesan_error = f"Akun Anda terdaftar sebagai {profile_role_display}, silakan gunakan login {role_display}!"
            else:
                login(request, user)
                LoginHistory.objects.create(user=user)
                return redirect('/')
        else:
            pesan_error = "Username atau password salah!"
            
    return render(request, 'inventory/login.html', {'error': pesan_error})

def halaman_register(request):
    if request.user.is_authenticated:
        return redirect('/')
        
    pesan_error = None
    if request.method == 'POST':
        u = request.POST.get('username')
        p1 = request.POST.get('password')
        p2 = request.POST.get('confirm_password')
        role = request.POST.get('role', 'KARYAWAN')
        
        if p1 != p2:
            pesan_error = "Password tidak cocok!"
        elif User.objects.filter(username=u).exists():
            pesan_error = "Username sudah terdaftar!"
        else:
            user = User.objects.create_user(username=u, password=p1)
            Profile.objects.create(user=user, role=role)
            login(request, user)
            LoginHistory.objects.create(user=user)
            return redirect('/')
            
    return render(request, 'inventory/register.html', {'error': pesan_error})

def proses_logout(request):
    logout(request)
    return redirect('halaman_login')

@login_required
def download_excel_stok(request):
    if Workbook is None:
        return HttpResponse("Library 'openpyxl' belum terinstall. Silakan jalankan: pip install openpyxl", status=500)

    # 1. Setup Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Gudang"
    
    # 2. Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Primary Blue
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # 3. Header Utama
    today_str = datetime.now().strftime("%d %B %Y")
    ws.merge_cells('A1:D1')
    ws['A1'] = f"LAPORAN STOK GUDANG - {today_str}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center_align

    # 4. Table Headers
    columns = ['Nama Bahan', 'Zona', 'Stok Sekarang', 'Satuan']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # 5. Data rows (Sorted by Zone)
    zona_order = {'TEA': 1, 'BAHAN_BAKU': 2, 'TOPPING': 3, 'DAIRY': 4}
    semua_bahan = sorted(BahanBaku.objects.all(), key=lambda x: zona_order.get(x.kategori_zona, 99))
    
    current_row = 4
    for bahan in semua_bahan:
        ws.cell(row=current_row, column=1, value=bahan.nama_bahan).border = border
        ws.cell(row=current_row, column=2, value=bahan.get_kategori_zona_display()).border = border
        ws.cell(row=current_row, column=3, value=bahan.stok_sekarang).border = border
        ws.cell(row=current_row, column=4, value=bahan.satuan).border = border
        current_row += 1

    # 6. Adjust Column Width
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    # 7. Response
    filename = f"Stok_Gudang_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ==============================================================================
# NATIVE REST API ENDPOINT: VOICE COMMAND INVENTORY MANAGER
# ==============================================================================
@csrf_exempt
def api_voice_command(request):
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Metode HTTP tidak diizinkan. Gunakan POST.'
        }, status=405)
    
    try:
        body = request.body.decode('utf-8')
        data = json.loads(body)
        
        command_text = data.get('command_text', '').lower().strip()
        petugas_id = data.get('petugas_id')
        
        if not command_text or not petugas_id:
            return JsonResponse({
                'status': 'error',
                'message': 'Parameter command_text dan petugas_id wajib diisi.'
            }, status=400)
            
        match = re.match(r'(tambah|kurang|pakai)\s+([a-z\s]+?)\s+(\d+)', command_text)
        
        if not match:
            return JsonResponse({
                'status': 'error',
                'message': 'Format perintah tidak dikenali. Contoh: "tambah boba 10"'
            }, status=400)
            
        aksi = match.group(1)
        nama_bahan = match.group(2).strip()
        jumlah = int(match.group(3))
        
        if jumlah <= 0:
            return JsonResponse({'status': 'error', 'message': 'Jumlah harus lebih besar dari 0.'}, status=400)

        with transaction.atomic():
            try:
                bahan = BahanBaku.objects.get(nama_bahan__iexact=nama_bahan)
            except BahanBaku.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Bahan baku "{nama_bahan}" tidak ditemukan.'
                }, status=404)
            
            stok_awal = bahan.stok_sekarang
            if aksi == 'tambah':
                stok_akhir = stok_awal + jumlah
            elif aksi in ['kurang', 'pakai']:
                stok_akhir = stok_awal - jumlah
                if stok_akhir < 0:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Stok tidak mencukupi. Stok "{nama_bahan}" saat ini {stok_awal}.'
                    }, status=400)
            else:
                return JsonResponse({'status': 'error', 'message': 'Aksi tidak valid.'}, status=400)
            
            bahan.stok_sekarang = stok_akhir
            bahan.save()
            
            user = User.objects.get(id=petugas_id)
            profile = getattr(user, 'profile', None)
            if profile and profile.role == 'MANAGER':
                return JsonResponse({'status': 'error', 'message': 'Akses ditolak: Manager tidak dapat memodifikasi stok.'}, status=403)
                
            RiwayatStok.objects.create(
                bahan=bahan,
                petugas=user,
                jumlah_baru=stok_akhir
            )
            
        return JsonResponse({
            'status': 'success',
            'message': f'Berhasil. Stok {nama_bahan} sekarang adalah {stok_akhir}.',
            'data': {
                'nama_bahan': bahan.nama_bahan,
                'stok_awal': stok_awal,
                'jumlah_perubahan': jumlah,
                'stok_akhir': stok_akhir,
            }
        }, status=200)

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Format JSON tidak valid.'}, status=400)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Petugas tidak ditemukan.'}, status=404)
    except DatabaseError:
        return JsonResponse({'status': 'error', 'message': 'Layanan Database tidak merespons.'}, status=503)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Internal Error: {str(e)}'}, status=500)


# ==============================================================================
# API ENDPOINTS UNTUK PEMANTAUAN & CATATAN MANAGER
# ==============================================================================

@login_required
def api_heartbeat(request):
    # Update activity timestamp
    profile, _ = Profile.objects.get_or_create(user=request.user)
    profile.last_activity = timezone.now()
    profile.save()
    
    # Deteksi apakah ada Manager lain/sama yang aktif memantau
    # (aktif jika last_activity kurang dari 15 detik yang lalu)
    now = timezone.now()
    active_managers = Profile.objects.filter(
        role='MANAGER',
        last_activity__gte=now - timezone.timedelta(seconds=15)
    )
    
    # Jika user saat ini adalah Manager, kita jangan hitung dia sendiri untuk dirinya sendiri,
    # tetapi untuk Karyawan dia tetap dianggap aktif memantau.
    manager_active = active_managers.exclude(user=request.user).exists() if profile.role == 'MANAGER' else active_managers.exists()
    
    # Ambil catatan aktif untuk Karyawan
    notes = []
    if profile.role == 'KARYAWAN':
        active_notes = CatatanManager.objects.filter(aktif=True).order_by('-waktu')
        for note in active_notes:
            notes.append({
                'id': note.id,
                'isi': note.isi,
                'waktu': note.waktu.strftime('%H:%M'),
                'pembuat': note.dibuat_oleh.username
            })
            
    return JsonResponse({
        'status': 'sukses',
        'manager_active': manager_active,
        'notes': notes
    })


@login_required
def tambah_catatan(request):
    profile = getattr(request.user, 'profile', None)
    if not profile or profile.role != 'MANAGER':
        return JsonResponse({'status': 'gagal', 'pesan': 'Akses ditolak: Hanya Manager yang dapat membuat catatan.'})
        
    if request.method == 'POST':
        isi = request.POST.get('isi', '').strip()
        if isi:
            CatatanManager.objects.create(isi=isi, dibuat_oleh=request.user)
            return JsonResponse({'status': 'sukses', 'pesan': 'Catatan berhasil dikirim.'})
        return JsonResponse({'status': 'gagal', 'pesan': 'Isi catatan tidak boleh kosong.'})
    return JsonResponse({'status': 'gagal', 'pesan': 'Method not allowed'})


@login_required
def nonaktifkan_catatan(request):
    if request.method == 'POST':
        note_id = request.POST.get('note_id')
        try:
            note = CatatanManager.objects.get(id=note_id)
            note.aktif = False
            note.save()
            return JsonResponse({'status': 'sukses'})
        except CatatanManager.DoesNotExist:
            return JsonResponse({'status': 'gagal', 'pesan': 'Catatan tidak ditemukan.'})
    return JsonResponse({'status': 'gagal', 'pesan': 'Method not allowed'})